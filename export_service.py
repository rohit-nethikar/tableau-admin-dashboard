"""
Export Service for Phase 4
Handles exporting dashboard data to CSV and Excel formats
"""

import csv
import json
import io
from datetime import datetime
from typing import Dict, Any, List


def export_metrics_to_csv(metrics: Dict[str, Any], filters_applied: Dict[str, str] = None) -> str:
    """
    Export overview metrics to CSV format
    Returns CSV string
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(['Tableau Admin Dashboard - Metrics Export'])
    writer.writerow(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    if filters_applied:
        writer.writerow(['Filters Applied:', json.dumps(filters_applied)])
    writer.writerow([])

    # Metrics section
    writer.writerow(['METRICS SUMMARY'])
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Workbooks', metrics.get('workbook_count', 0)])
    writer.writerow(['Data Sources', metrics.get('datasource_count', 0)])
    writer.writerow(['Stale Items', metrics.get('stale_count', 0)])
    writer.writerow(['Custom Views', metrics.get('custom_view_count', 0)])
    writer.writerow(['Subscriptions', metrics.get('subscription_count', 0)])
    writer.writerow(['Users', metrics.get('user_count', 0)])
    writer.writerow(['Average Health Score', metrics.get('avg_score', 0)])
    writer.writerow([])

    # Severity breakdown
    writer.writerow(['FINDINGS BY SEVERITY'])
    writer.writerow(['Severity', 'Count'])
    severity_counts = metrics.get('severity_counts', {})
    writer.writerow(['Critical', severity_counts.get('critical', 0)])
    writer.writerow(['High', severity_counts.get('high', 0)])
    writer.writerow(['Medium', severity_counts.get('medium', 0)])
    writer.writerow(['Low', severity_counts.get('low', 0)])
    writer.writerow([])

    # Health score distribution
    writer.writerow(['HEALTH SCORE DISTRIBUTION'])
    writer.writerow(['Category', 'Count'])
    score_buckets = metrics.get('score_buckets', {})
    writer.writerow(['Good (80+)', score_buckets.get('good', 0)])
    writer.writerow(['Warning (50-79)', score_buckets.get('warning', 0)])
    writer.writerow(['Critical (<50)', score_buckets.get('critical', 0)])
    writer.writerow([])

    # User roles
    writer.writerow(['USER DISTRIBUTION BY ROLE'])
    writer.writerow(['Role', 'Count'])
    user_roles = metrics.get('user_roles', {})
    for role, count in user_roles.items():
        writer.writerow([role, count])
    writer.writerow([])

    # Content type breakdown
    writer.writerow(['CONTENT TYPE BREAKDOWN'])
    writer.writerow(['Type', 'Count'])
    content_types = metrics.get('content_by_type', {})
    for content_type, count in content_types.items():
        writer.writerow([content_type, count])

    return output.getvalue()


def export_findings_to_csv(findings: List[Dict[str, Any]]) -> str:
    """
    Export findings/alerts to CSV format
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(['Tableau Admin Dashboard - Findings Export'])
    writer.writerow(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])

    # Findings
    writer.writerow(['ID', 'Title', 'Resource', 'Severity', 'Status', 'Description'])
    for finding in findings:
        writer.writerow([
            finding.get('id', ''),
            finding.get('title', ''),
            finding.get('resource_name', ''),
            finding.get('severity', ''),
            finding.get('status', ''),
            finding.get('description', '')
        ])

    return output.getvalue()


def export_to_excel(metrics: Dict[str, Any], findings: List[Dict[str, Any]] = None) -> bytes:
    """
    Export to Excel format with multiple sheets
    Returns Excel file as bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl not available
        print("Warning: openpyxl not installed, falling back to CSV")
        return export_metrics_to_csv(metrics).encode('utf-8')

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Remove default sheet

    # Sheet 1: Overview
    ws = workbook.create_sheet('Overview')
    ws['A1'] = 'Tableau Admin Dashboard - Metrics Export'
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

    row = 4
    ws[f'A{row}'] = 'METRICS SUMMARY'
    row += 1
    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Value'
    row += 1

    metrics_data = [
        ('Workbooks', metrics.get('workbook_count', 0)),
        ('Data Sources', metrics.get('datasource_count', 0)),
        ('Stale Items', metrics.get('stale_count', 0)),
        ('Custom Views', metrics.get('custom_view_count', 0)),
        ('Subscriptions', metrics.get('subscription_count', 0)),
        ('Users', metrics.get('user_count', 0)),
        ('Average Health Score', metrics.get('avg_score', 0)),
    ]

    for metric_name, value in metrics_data:
        ws[f'A{row}'] = metric_name
        ws[f'B{row}'] = value
        row += 1

    # Sheet 2: Findings
    if findings:
        ws_findings = workbook.create_sheet('Findings')
        ws_findings['A1'] = 'Findings Report'
        ws_findings['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

        row = 4
        headers = ['ID', 'Title', 'Resource', 'Severity', 'Status']
        for col, header in enumerate(headers, 1):
            ws_findings.cell(row=row, column=col, value=header)

        row += 1
        for finding in findings:
            ws_findings.cell(row=row, column=1, value=finding.get('id', ''))
            ws_findings.cell(row=row, column=2, value=finding.get('title', ''))
            ws_findings.cell(row=row, column=3, value=finding.get('resource_name', ''))
            ws_findings.cell(row=row, column=4, value=finding.get('severity', ''))
            ws_findings.cell(row=row, column=5, value=finding.get('status', ''))
            row += 1

    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def export_filters_to_json(filters: Dict[str, Any]) -> str:
    """Export current filters as JSON"""
    return json.dumps({
        'exported_at': datetime.now().isoformat(),
        'filters': filters
    }, indent=2)


def format_export_filename(export_type: str) -> str:
    """Generate a filename for export based on type and timestamp"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"tableau_dashboard_{export_type}_{timestamp}"
